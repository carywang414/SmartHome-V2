import cv2
import time
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from face_recognition import build_white_list_embeddings, recognize_face

# === 初始化資料夾與 Excel ===
history_folder = "history"
record_file = os.path.join(history_folder, "access_log.xlsx")
os.makedirs(history_folder, exist_ok=True)

if not os.path.exists(record_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["日期", "時間", "姓名"])
    wb.save(record_file)
else:
    wb = load_workbook(record_file)
    ws = wb.active

def log_access(name):
    now = datetime.now()
    date_str = now.strftime("%Y%m%d")
    time_str = now.strftime("%H%M")
    ws.append([date_str, time_str, name])
    wb.save(record_file)

# === 建立白名單資料庫 ===
print("🔍 建立允許通行的 embedding 資料庫...")
database = build_white_list_embeddings('face')

if not database:
    print("❌ 沒有找到有效的人臉資料，請將照片放入 face/ 資料夾")
    exit()

print("✅ 資料庫載入完成，啟動攝影機辨識中...")
cap = cv2.VideoCapture(0)

recognized_name = None
start_time = None

final_frame = None  # 保留最後畫面用於 unknown 截圖

while True:
    ret, frame = cap.read()
    if not ret:
        break

    name, dist = recognize_face(frame, database)
    final_frame = frame.copy()  # 儲存目前畫面

    if name not in ["No face"]:
        current_required = 10 if name == "Unknown" else 3
        if recognized_name == name:
            elapsed = time.time() - start_time
            if name == "Unknown":
                label = f"❌ no famil detected {elapsed:.1f} s"
                color = (0, 0, 255)
            else:
                label = f"✅ {name} my famil detected {elapsed:.1f} s"
                color = (0, 255, 0)
            if elapsed >= current_required:
                print(f"🎉 {name} 持續存在 {current_required} 秒，自動關閉攝影機")
                log_access(name)

                # 如果是 unknown，儲存截圖
                if name == "Unknown":
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    img_path = os.path.join(history_folder, f"unknown_{timestamp}.png")
                    cv2.imwrite(img_path, final_frame)
                    print(f"📸 Unknown 截圖已儲存：{img_path}")

                break
        else:
            recognized_name = name
            start_time = time.time()
            if name == "Unknown":
                label = "❌ no famil（開始計時）"
                color = (0, 0, 255)
            else:
                label = f"✅ {name} 允許進入（開始計時）"
                color = (0, 255, 0)
    else:
        # 沒臉重置計時
        recognized_name = None
        start_time = None
        label = "no face"
        color = (128, 128, 128)

    cv2.putText(frame, label, (30, 40), cv2.FONT_HERSHEY_SIMPLEX, 1, color, 2)
    cv2.imshow("Access Control System", frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        print("👋 手動結束辨識")
        break

cap.release()
cv2.destroyAllWindows()